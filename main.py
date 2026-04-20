from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from selenium.webdriver import ChromeOptions, EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait


DEFAULT_LOGIN_URL = (
    "https://authtest.mynagad.com:10900/authentication-service-provider-1.0/login"
)
DEFAULT_SYSTEM_URL = "https://systest.mynagad.com:20020/ui/system/#/home"
DEFAULT_CREATE_ROUTE = "/ui/system/#/fee-commission-management/create/biller-merchant"
DEFAULT_CREDENTIAL_FILES = ("credentials.txt", "credentials")
DEFAULT_TIMEOUT = 30
SLAB_MIN = "0"
SLAB_MAX = "9999999999"
PLAYER_LABELS = {
    "UDD": "UDDOKTA",
    "DH": "DISTRIBUTOR",
    "AD": "ADVANCE COMMISSION",
}


@dataclass
class Credentials:
    username: str
    password: str


@dataclass
class WorkflowConfig:
    login_url: str
    system_url: str
    create_url: str
    workbook_path: Path
    credentials_path: Path
    browser: str
    headless: bool
    merchant_account_no: str
    service_query: str
    effective_time_hh24: str
    timeout: int
    failure_artifacts_dir: Path


@dataclass
class ChargeEntry:
    channel: str
    charging_type: str
    player: str
    value: str
    value_type: str = "RATE"
    minimum: str = ""
    maximum: str = ""


@dataclass
class PayeeWorkflow:
    payee: str
    fee_entries: list[ChargeEntry]
    commission_entries: list[ChargeEntry]


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def stringify_numeric(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.9f}".rstrip("0").rstrip(".")
    return str(value).strip()


def parse_credentials(path: Path) -> Credentials:
    if not path.exists():
        raise FileNotFoundError(f"Credentials file not found: {path}")

    username = ""
    password = ""
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = normalize(key)
        if key == "USERNAME":
            username = value.strip()
        elif key == "PASSWORD":
            password = value.strip()

    if not username or not password:
        raise ValueError(
            "Credentials file must contain 'username: ...' and 'password: ...' lines."
        )
    return Credentials(username=username, password=password)


def build_create_url(system_url: str) -> str:
    parts = urlsplit(system_url)
    return urlunsplit((parts.scheme, parts.netloc, DEFAULT_CREATE_ROUTE, "", ""))


def detect_workbook(explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path

    workbooks = sorted(
        [
            *Path.cwd().glob("*.xlsx"),
            *Path.cwd().glob("*.xlsm"),
        ]
    )
    if not workbooks:
        raise FileNotFoundError(
            "No .xlsx workbook found in the current folder. "
            "Place the Excel file here or pass --workbook."
        )
    return workbooks[0].resolve()


def detect_credentials(explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Credentials file not found: {path}")
        return path

    for name in DEFAULT_CREDENTIAL_FILES:
        path = Path.cwd() / name
        if path.exists():
            return path.resolve()

    expected = " or ".join(DEFAULT_CREDENTIAL_FILES)
    raise FileNotFoundError(
        f"Credentials file not found. Create {expected} in this folder or pass --credentials."
    )


def parse_workbook(path: Path) -> list[PayeeWorkflow]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    parsed_rows: dict[str, dict[str, dict[str, str]]] = {}
    service_fees: dict[str, dict[str, str]] = {}

    for index in range(len(rows) - 1):
        header_row = rows[index]
        value_row = rows[index + 1]
        cells = [normalize(cell) for cell in header_row]
        if "APP" not in cells and "USSD" not in cells:
            continue
        if "UDDOKTA" not in cells and "CUSTOMER" not in cells:
            continue

        channel_idx = next(
            (i for i, cell in enumerate(cells) if cell in {"APP", "USSD"}), None
        )
        payee_idx = next(
            (i for i, cell in enumerate(cells) if cell in {"UDDOKTA", "CUSTOMER"}),
            None,
        )
        if channel_idx is None or payee_idx is None or payee_idx <= channel_idx:
            continue

        channel = cells[channel_idx]
        payee = cells[payee_idx]
        value_map: dict[str, str] = {}
        service_fee = ""

        for col in range(payee_idx + 1, len(header_row)):
            header = normalize(header_row[col])
            value = stringify_numeric(value_row[col] if col < len(value_row) else None)
            if not header or not value:
                continue
            if header == "SERVICE FEE":
                service_fee = value
                continue
            value_map[header] = value

        if value_map:
            parsed_rows.setdefault(payee, {})[channel] = value_map
        if service_fee:
            service_fees.setdefault(payee, {})[channel] = service_fee

    workflows: list[PayeeWorkflow] = []
    for payee in ("UDDOKTA", "CUSTOMER"):
        payee_rows = parsed_rows.get(payee, {})
        if not payee_rows:
            continue

        fee_entries: list[ChargeEntry] = []
        commission_entries: list[ChargeEntry] = []
        for channel in ("APP", "USSD"):
            fee_value = service_fees.get(payee, {}).get(channel)
            if fee_value:
                fee_entries.append(
                    ChargeEntry(
                        channel=channel,
                        charging_type="FEE",
                        player=payee,
                        value=fee_value,
                    )
                )

            for raw_player, value in payee_rows.get(channel, {}).items():
                player = PLAYER_LABELS.get(raw_player, raw_player)
                commission_entries.append(
                    ChargeEntry(
                        channel=channel,
                        charging_type="COMMISSION",
                        player=player,
                        value=value,
                    )
                )

        workflows.append(
            PayeeWorkflow(
                payee=payee,
                fee_entries=fee_entries,
                commission_entries=commission_entries,
            )
        )

    if not workflows:
        raise ValueError(
            "Could not find APP/USSD fee commission rows in the workbook's first sheet."
        )
    return workflows


def build_driver(browser: str, headless: bool) -> WebDriver:
    browser_name = browser.lower()
    if browser_name == "chrome":
        options = ChromeOptions()
        if headless:
            options.add_argument("--headless=new")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        return webdriver.Chrome(options=options)

    options = EdgeOptions()
    options.use_chromium = True
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    return webdriver.Edge(options=options)


class PortalAutomation:
    def __init__(self, driver: WebDriver, timeout: int) -> None:
        self.driver = driver
        self.wait = WebDriverWait(driver, timeout)
        self.current_step = "initializing"

    def xpath_literal(self, value: str) -> str:
        if "'" not in value:
            return f"'{value}'"
        if '"' not in value:
            return f'"{value}"'
        parts = value.split("'")
        quoted = ", \"'\", ".join(f"'{part}'" for part in parts)
        return f"concat({quoted})"

    def wait_visible(self, xpath: str):
        return self.wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))

    def wait_clickable(self, xpath: str):
        return self.wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))

    def click(self, xpath: str) -> None:
        attempts = 3
        for attempt in range(attempts):
            try:
                element = self.wait_clickable(xpath)
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});", element
                )
                try:
                    element.click()
                except Exception:
                    self.driver.execute_script("arguments[0].click();", element)
                return
            except StaleElementReferenceException:
                if attempt == attempts - 1:
                    raise

    def set_input(self, xpath: str, value: str, clear: bool = True) -> None:
        field = self.wait_visible(xpath)
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", field)
        field.click()
        if clear:
            field.send_keys(Keys.CONTROL, "a")
            field.send_keys(Keys.DELETE)
        field.send_keys(value)

    def click_button(self, text: str) -> None:
        text_lower = text.lower()
        literal = self.xpath_literal(text_lower)
        xpath = (
            f"//button[translate(normalize-space(), "
            f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')={literal}] | "
            f"//a[translate(normalize-space(), "
            f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')={literal}] | "
            f"//span[translate(normalize-space(), "
            f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')={literal}] | "
            f"//span[translate(normalize-space(), "
            f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')={literal}]/ancestor::button[1]"
        )
        self.click(xpath)

    def set_field_after_label(self, label: str, value: str) -> None:
        literal = self.xpath_literal(label)
        xpath = (
            f"(//*[contains(normalize-space(), {literal})]"
            f"/following::*[self::input or self::textarea][1])[1]"
        )
        self.set_input(xpath, value)

    def open_dropdown_after_label(self, label: str) -> None:
        literal = self.xpath_literal(label)
        xpath = (
            f"(//*[contains(normalize-space(), {literal})]"
            f"/following::*[(self::input or @role='combobox' or self::div or self::span)"
            f" and not(self::textarea)][1])[1]"
        )
        self.click(xpath)

    def find_select_after_label(self, label: str):
        literal = self.xpath_literal(label)
        xpath = (
            f"(//*[contains(normalize-space(), {literal})]"
            f"/following::*[self::select][1])[1]"
        )
        return self.wait_visible(xpath)

    def choose_dropdown_option(self, option_text: str) -> None:
        literal = self.xpath_literal(option_text)
        option_xpath = (
            f"//*[self::li or self::div or self::span]"
            f"[normalize-space()={literal} or contains(normalize-space(), {literal})]"
        )
        self.click(option_xpath)

    def choose_radio_option(self, option_text: str) -> None:
        literal = self.xpath_literal(option_text)
        xpath = (
            f"//label[contains(normalize-space(), {literal})] | "
            f"//*[contains(normalize-space(), {literal})]/preceding::input[@type='radio'][1] | "
            f"//*[contains(normalize-space(), {literal})]/following::input[@type='radio'][1]"
        )
        self.click(xpath)

    def select_dropdown_value(self, label: str, value: str) -> None:
        try:
            select_element = self.find_select_after_label(label)
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});", select_element
            )
            try:
                select_element.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", select_element)
            select = Select(select_element)
            try:
                select.select_by_visible_text(value)
            except Exception:
                target = normalize(value)
                matched = False
                for option in select.options:
                    option_text = normalize(option.text)
                    if target and (option_text == target or target in option_text):
                        option_value = option.get_attribute("value")
                        if option_value is not None:
                            select.select_by_value(option_value)
                        else:
                            option.click()
                        self.driver.execute_script(
                            """
                            const select = arguments[0];
                            select.dispatchEvent(new Event('input', { bubbles: true }));
                            select.dispatchEvent(new Event('change', { bubbles: true }));
                            """,
                            select_element,
                        )
                        matched = True
                        break
                if not matched:
                    raise
            return
        except Exception:
            pass

        self.open_dropdown_after_label(label)
        time.sleep(0.4)
        self.choose_dropdown_option(value)

    def wait_for_any_text(self, texts: Iterable[str]) -> None:
        conditions = [
            f"contains(normalize-space(), {self.xpath_literal(text)})" for text in texts
        ]
        xpath = "//*[" + " or ".join(conditions) + "]"
        self.wait_visible(xpath)

    def get_pending_approval_message(self) -> str:
        candidates = self.driver.find_elements(
            By.XPATH,
            "//*[contains(normalize-space(), 'Approval is pending') or "
            "contains(normalize-space(), 'approve pending item first')]",
        )
        for element in candidates:
            text = " ".join(element.text.split())
            if text:
                return text.strip("' ")
        return ""

    def login(self, login_url: str, username: str, password: str) -> None:
        self.current_step = "login"
        self.driver.get(login_url)
        self.set_input(
            "(//input[@type='text' or @type='email' or contains(@placeholder, 'user')])[1]",
            username,
        )
        self.set_input("(//input[@type='password'])[1]", password)
        self.click_button("LOG IN")
        self.wait_for_any_text(["Campaign Management", "Fee Commission Management", "Home"])

    def go_to_create_page(self, create_url: str) -> None:
        self.current_step = "open create page"
        self.driver.get(create_url)
        self.wait_for_any_text(["Biller Fee Commission Create", "Fee Commission For"])

    def search_and_select_merchant(self, merchant_account_no: str) -> None:
        self.current_step = f"search merchant {merchant_account_no}"
        self.click(
            "(//label[contains(normalize-space(), 'Merchant')]/following::div[contains(@class, 'input-group')][1]"
            "//span[contains(@class, 'btn') or contains(@class, 'input-group-text')])[1]"
        )
        self.wait_for_any_text(["Search Merchant", "Mobile Account No."])
        self.set_field_after_label("Mobile Account No.", merchant_account_no)
        self.current_step = f"click search for merchant {merchant_account_no}"
        self.click("//ngb-modal-window//button[contains(@class, 'searchButton') and contains(normalize-space(), 'Search')]")
        account_literal = self.xpath_literal(merchant_account_no)
        self.wait.until(
            lambda d: d.find_elements(
                By.XPATH,
                f"//ngb-modal-window//tr[td[normalize-space()={account_literal}]]",
            )
        )
        row_xpath = f"//ngb-modal-window//tr[td[normalize-space()={account_literal}]]"
        radio_xpath = f"{row_xpath}//input[@type='radio']"
        try:
            self.current_step = f"select merchant radio for {merchant_account_no}"
            for attempt in range(3):
                try:
                    row = self.wait_visible(row_xpath)
                    self.driver.execute_script(
                        "arguments[0].scrollIntoView({block: 'center'});", row
                    )
                    radio = self.wait_visible(radio_xpath)
                    radio_cell = row.find_element(By.XPATH, ".//td[input[@type='radio']]")
                    try:
                        ActionChains(self.driver).move_to_element(radio_cell).click().perform()
                    except Exception:
                        pass
                    if not radio.is_selected():
                        try:
                            ActionChains(self.driver).move_to_element(radio).click().perform()
                        except Exception:
                            pass
                    if not radio.is_selected():
                        self.driver.execute_script("arguments[0].click();", radio)
                    if not radio.is_selected():
                        radio.send_keys(Keys.SPACE)
                    self.wait.until(lambda d: self.wait_visible(radio_xpath).is_selected())
                    break
                except StaleElementReferenceException:
                    if attempt == 2:
                        raise
        except TimeoutException:
            pass
        self.current_step = f"confirm selected merchant {merchant_account_no}"
        self.click("//span[@id='btnSubmitId' and contains(normalize-space(), 'Select')]")
        self.wait.until(
            lambda d: not d.find_elements(By.XPATH, "//ngb-modal-window[contains(@class, 'show')]")
        )
        merchant_input = self.wait_visible("//input[@name='selectedPartnerLabel']")
        self.wait.until(lambda d: merchant_input.get_attribute("value").strip() != "")

    def select_service(self, service_query: str) -> None:
        self.current_step = f"select service {service_query}"
        self.select_dropdown_value("Service", service_query)

    def create_payee_header(self, merchant_account_no: str, service_query: str, payee: str) -> None:
        self.current_step = f"create header for payee {payee}"
        self.search_and_select_merchant(merchant_account_no)
        self.select_service(service_query)
        self.select_dropdown_value("Payee", payee)
        approval_message = self.get_pending_approval_message()
        if approval_message:
            raise RuntimeError(approval_message)
        self.click_button("Submit")
        self.wait_for_any_text(["Effective Date", "ADD FEE COMMISSION"])

    def fill_effective_time(self, effective_time_hh24: str) -> None:
        self.current_step = f"fill effective time {effective_time_hh24}"
        self.set_field_after_label("Time (HH24)", effective_time_hh24)

    def fill_common_entry_fields(self, entry: ChargeEntry) -> None:
        self.current_step = (
            f"fill entry channel={entry.channel} type={entry.charging_type} "
            f"player={entry.player} value={entry.value}"
        )
        self.set_field_after_label("Slab Min.", SLAB_MIN)
        self.set_field_after_label("Slab Max.", SLAB_MAX)
        self.click_button("More")
        self.select_dropdown_value("Channel", entry.channel)
        self.choose_radio_option(entry.charging_type)
        self.select_dropdown_value("Player", entry.player)
        self.choose_radio_option(entry.value_type)
        self.set_field_after_label("Value", entry.value)
        if entry.minimum:
            self.set_field_after_label("Minimum", entry.minimum)
        if entry.maximum:
            self.set_field_after_label("Maximum", entry.maximum)

    def add_charge_entry(self, entry: ChargeEntry) -> None:
        self.current_step = (
            f"add entry channel={entry.channel} type={entry.charging_type} "
            f"player={entry.player}"
        )
        self.fill_common_entry_fields(entry)
        self.click_button("Add new")
        time.sleep(1.0)

    def register(self) -> None:
        self.current_step = "register workflow"
        self.click_button("Register")
        self.wait_for_any_text(["registered successfully", "Search Biller Merchant Fee Commission"])

    def capture_failure_artifacts(self, output_dir: Path) -> list[Path]:
        artifacts: list[Path] = []
        output_dir.mkdir(parents=True, exist_ok=True)

        screenshot_path = output_dir / "failure.png"
        page_source_path = output_dir / "page_source.html"

        try:
            if self.driver.save_screenshot(str(screenshot_path)):
                artifacts.append(screenshot_path)
        except Exception:
            pass

        try:
            page_source_path.write_text(self.driver.page_source, encoding="utf-8")
            artifacts.append(page_source_path)
        except Exception:
            pass

        return artifacts


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Automate Nagad fee commission setup from the screenshot-defined workflow."
    )
    parser.add_argument("--workbook", help="Path to the Excel workbook.")
    parser.add_argument("--credentials", help="Path to credentials.txt.")
    parser.add_argument("--login-url", default=DEFAULT_LOGIN_URL)
    parser.add_argument("--system-url", default=DEFAULT_SYSTEM_URL)
    parser.add_argument("--browser", choices=["edge", "chrome"], default="edge")
    parser.add_argument("--merchant-account-no", required=True)
    parser.add_argument(
        "--service-query",
        required=True,
        help="Text to type into the service dropdown, for example 'asfia' or 'asfia | 1333'.",
    )
    parser.add_argument(
        "--effective-time-hh24",
        required=True,
        help="Time to enter in the Time (HH24) field.",
    )
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT)
    parser.add_argument("--headless", action="store_true")
    return parser


def run(config: WorkflowConfig) -> None:
    credentials = parse_credentials(config.credentials_path)
    workflows = parse_workbook(config.workbook_path)
    driver = build_driver(config.browser, config.headless)

    try:
        portal = PortalAutomation(driver, config.timeout)
        portal.login(config.login_url, credentials.username, credentials.password)

        for workflow in workflows:
            portal.go_to_create_page(config.create_url)
            portal.create_payee_header(
                merchant_account_no=config.merchant_account_no,
                service_query=config.service_query,
                payee=workflow.payee,
            )
            portal.fill_effective_time(config.effective_time_hh24)

            for entry in workflow.fee_entries:
                portal.add_charge_entry(entry)

            for entry in workflow.commission_entries:
                portal.add_charge_entry(entry)

            portal.register()
    except Exception as exc:
        details = [f"Automation failed during step: {portal.current_step}"]
        artifacts = portal.capture_failure_artifacts(config.failure_artifacts_dir)
        if artifacts:
            joined = ", ".join(str(path) for path in artifacts)
            details.append(f"Artifacts saved: {joined}")
        raise RuntimeError(f"{'; '.join(details)}; Root cause: {exc!r}") from exc
    finally:
        driver.quit()


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    try:
        workbook_path = detect_workbook(args.workbook)
        credentials_path = detect_credentials(args.credentials)
        create_url = build_create_url(args.system_url)
        config = WorkflowConfig(
            login_url=args.login_url,
            system_url=args.system_url,
            create_url=create_url,
            workbook_path=workbook_path,
            credentials_path=credentials_path,
            browser=args.browser,
            headless=args.headless,
            merchant_account_no=args.merchant_account_no,
            service_query=args.service_query,
            effective_time_hh24=args.effective_time_hh24,
            timeout=args.timeout,
            failure_artifacts_dir=Path.cwd() / "artifacts",
        )
        run(config)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
